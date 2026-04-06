print("🚨 SERVER NOVO CARREGADO 🚨")
import cloudinary
import cloudinary.uploader
import os
import io
import json
import uuid
import jwt
import bcrypt
import zipfile
import requests
import logging
import base64
import numpy as np
import cv2
import pytesseract

from pdf2image import convert_from_bytes
from openpyxl.styles import PatternFill

from pathlib import Path
from datetime import datetime, timedelta
from typing import List, Optional

import aiofiles
from dotenv import load_dotenv
from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse, StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr
from openpyxl import Workbook

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

cloudinary.config(
    cloud_name="dkfpglash",
    api_key="489862842735213",
    api_secret="D8e_IQZMDtRpq3FSpXCjMySIqBA",
    secure=True
)

print("CLOUDINARY_URL exists:", bool(os.getenv("CLOUDINARY_URL")))
print("CLOUDINARY_URL value:", os.getenv("CLOUDINARY_URL"))
print("Cloudinary config:", cloudinary.config().cloud_name)

mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get("DB_NAME", "coreme_uepa")]

JWT_SECRET = os.environ.get("JWT_SECRET", "uepa_coreme_secret_key_2025")
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

UPLOAD_DIR = ROOT_DIR / "uploads" / "pdfs"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(
    title="COREME UEPA - Sistema de Envio de PDF",
    docs_url="/docs",
    redoc_url="/redoc",
    openapi_url="/openapi.json"
)
from fastapi.middleware.cors import CORSMiddleware

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:8081",
        "http://127.0.0.1:8081",
        "http://localhost:19006",
        "http://127.0.0.1:19006",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
@app.api_route("/", methods=["GET", "HEAD"])
async def root_test():
    return {"ok": True, "message": "root funcionando"}

@app.get("/ping")
async def ping():
    return {"ping": "pong"}
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


class UserBase(BaseModel):
    email: EmailStr
    full_name: str
    program: str
    year: Optional[str] = None
    role: str = "resident"
    cpf: Optional[str] = None
    scenario: Optional[str] = None


class UserCreate(BaseModel):
    email: EmailStr
    password: str
    full_name: str
    program: str
    year: Optional[str] = None
    role: str
    cpf: Optional[str] = None
    scenario: Optional[str] = None


class UserLogin(BaseModel):
    email: EmailStr
    password: str


class User(UserBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=datetime.utcnow)
    is_active: bool = True


class UserResponse(BaseModel):
    id: str
    email: str
    full_name: str
    program: str
    year: Optional[str] = None
    role: str
    cpf: Optional[str] = None
    scenario: Optional[str] = None
    created_at: datetime
    is_active: bool


class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse


class Submission(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_name: str
    user_email: str
    program: str
    year: str
    reference_month: str
    reference_month_name: str

    file_path: str
    file_name: str

    aulas_file_path: Optional[str] = None
    aulas_file_name: Optional[str] = None

    orientacao_file_path: Optional[str] = None
    orientacao_file_name: Optional[str] = None

    submitted_at: datetime = Field(default_factory=datetime.utcnow)


class SubmissionResponse(BaseModel):
    id: str
    user_id: str
    user_name: str
    user_email: str
    program: str
    year: Optional[str] = None
    reference_month: str
    reference_month_name: str
    submitted_at: datetime
    status: str = "Enviado"


class SystemConfig(BaseModel):
    id: str = "system_config"
    logo_base64: Optional[str] = None
    system_name: str = "COREME UEPA"
    primary_color: str = "#1E3A8A"
    secondary_color: str = "#FFFFFF"
    login_image_base64: Optional[str] = None
    institutional_message: str = "Sistema de Envio Mensal de Documentos da Residência Médica"
    updated_at: datetime = Field(default_factory=datetime.utcnow)


class SystemConfigUpdate(BaseModel):
    logo_base64: Optional[str] = None
    system_name: Optional[str] = None
    primary_color: Optional[str] = None
    secondary_color: Optional[str] = None
    login_image_base64: Optional[str] = None
    institutional_message: Optional[str] = None


MONTH_NAMES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
}


def get_drive_service():
    creds_json = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")
    if not creds_json:
        raise Exception("Credenciais do Google não configuradas")

    info = json.loads(creds_json)
    credentials = service_account.Credentials.from_service_account_info(
        info,
        scopes=["https://www.googleapis.com/auth/drive"]
    )
    return build("drive", "v3", credentials=credentials)


async def upload_to_drive(file_content: bytes, filename: str, mime_type: Optional[str]):
    service = get_drive_service()
    folder_id = os.getenv("GOOGLE_DRIVE_FOLDER_ID")

    if not folder_id:
        raise Exception("GOOGLE_DRIVE_FOLDER_ID não configurado")

    file_stream = io.BytesIO(file_content)
    file_metadata = {
        "name": filename,
        "parents": [folder_id]
    }

    media = MediaIoBaseUpload(
        file_stream,
        mimetype=mime_type or "application/pdf",
        resumable=True
    )

    created_file = service.files().create(
        body=file_metadata,
        media_body=media,
        fields="id,name"
    ).execute()

    return created_file


def download_from_drive(file_id: str) -> io.BytesIO:
    service = get_drive_service()
    request = service.files().get_media(fileId=file_id)
    buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(buffer, request)

    done = False
    while not done:
        _, done = downloader.next_chunk()

    buffer.seek(0)
    return buffer


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode("utf-8"), hashed.encode("utf-8"))


def create_token(user_id: str, role: str) -> str:
    expiration = datetime.utcnow() + timedelta(hours=JWT_EXPIRATION_HOURS)
    payload = {
        "sub": user_id,
        "role": role,
        "exp": expiration
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def decode_token(token: str) -> dict:
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")


async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    payload = decode_token(credentials.credentials)
    user = await db.users.find_one({"id": payload["sub"]})
    if not user:
        raise HTTPException(status_code=401, detail="Usuário não encontrado")
    return user


async def get_admin_user(current_user: dict = Depends(get_current_user)) -> dict:
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Acesso negado. Apenas administradores.")
    return current_user


def get_reference_month_info() -> dict:
    now = datetime.utcnow()
    current_day = now.day

    if now.month == 1:
        ref_year = now.year - 1
        ref_month = 12
    else:
        ref_year = now.year
        ref_month = now.month - 1

    reference_month = f"{ref_year}-{ref_month:02d}"
    reference_month_name = f"{MONTH_NAMES_PT[ref_month]} {ref_year}"

    deadline_day = 4
    
    FORCE_UPLOAD_OPEN = True
    is_within_deadline = FORCE_UPLOAD_OPEN or (current_day <= deadline_day)

    if is_within_deadline:
        deadline = datetime(now.year, now.month, deadline_day, 23, 59, 59)
    else:
        deadline = None

    return {
        "reference_month": reference_month,
        "reference_month_name": reference_month_name,
        "deadline": deadline.isoformat() if deadline else None,
        "is_within_deadline": is_within_deadline,
        "current_day": current_day,
        "deadline_day": deadline_day,
        "message": "Prazo de envio encerrado." if not is_within_deadline else f"Prazo: até dia {deadline_day} às 23:59"
    }

def sanitize_filename(name: str) -> str:
    import re
    import unicodedata

    name = unicodedata.normalize("NFKD", name).encode("ASCII", "ignore").decode("ASCII")
    name = re.sub(r"[^\w\s-]", "", name)
    name = re.sub(r"[\s]+", "_", name)
    return name.lower()


def get_user_type_folder(user: dict) -> str:
    role = user.get("role", "").lower()
    if role == "preceptor":
        return "preceptores"
    return "residentes"


def generate_frequencia_folder(user: dict, reference_month: str) -> str:
    user_type = get_user_type_folder(user)
    return f"coreme/{user_type}/frequencia/{reference_month}"


def generate_user_material_folder(user: dict, reference_month: str) -> str:
    user_type = get_user_type_folder(user)
    program = sanitize_filename(user.get("program", "sem_programa"))
    full_name = sanitize_filename(user.get("full_name", "sem_nome"))
    return f"coreme/{user_type}/{program}/{full_name}/{reference_month}"


def upload_to_cloudinary(file_content: bytes, filename: str, folder: str, public_id: str):
    print("Enviando para Cloudinary...")

    result = cloudinary.uploader.upload(
        file_content,
        resource_type="raw",
        folder=folder,
        public_id=public_id,
        display_name=filename,
        overwrite=True
    )

    print("Upload concluído")
    return result


TESSERACT_CMD = os.getenv("TESSERACT_CMD", "tesseract")
pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

MONTH_TOKENS_PT = {
    1: ["janeiro", "jan"],
    2: ["fevereiro", "fev"],
    3: ["marco", "março", "mar"],
    4: ["abril", "abr"],
    5: ["maio", "mai"],
    6: ["junho", "jun"],
    7: ["julho", "jul"],
    8: ["agosto", "ago"],
    9: ["setembro", "set"],
    10: ["outubro", "out"],
    11: ["novembro", "nov"],
    12: ["dezembro", "dez"],
}


def normalize_ocr_text(text: str) -> str:
    if not text:
        return ""

    text = text.lower()
    text = text.replace("ç", "c")
    text = text.replace("á", "a").replace("à", "a").replace("ã", "a").replace("â", "a")
    text = text.replace("é", "e").replace("ê", "e")
    text = text.replace("í", "i")
    text = text.replace("ó", "o").replace("ô", "o").replace("õ", "o")
    text = text.replace("ú", "u")
    return text


def month_found_in_text(text: str, expected_month: int) -> bool:
    clean = normalize_ocr_text(text)
    tokens = MONTH_TOKENS_PT.get(expected_month, [])
    return any(token in clean for token in tokens)


def pil_to_cv2(pil_image):
    img = np.array(pil_image)
    if len(img.shape) == 2:
        return img
    return cv2.cvtColor(img, cv2.COLOR_RGB2BGR)


def crop_relative_region(img, box):
    """
    box = (x1, y1, x2, y2) em proporção da imagem
    """
    h, w = img.shape[:2]
    x1, y1, x2, y2 = box

    x1 = max(0, min(w, int(x1 * w)))
    x2 = max(0, min(w, int(x2 * w)))
    y1 = max(0, min(h, int(y1 * h)))
    y2 = max(0, min(h, int(y2 * h)))

    return img[y1:y2, x1:x2]


def detect_handwritten_mark(region_img, min_pixels=250):
    if region_img is None or region_img.size == 0:
        return False, 0

    if len(region_img.shape) == 3:
        gray = cv2.cvtColor(region_img, cv2.COLOR_BGR2GRAY)
    else:
        gray = region_img.copy()

    gray = cv2.GaussianBlur(gray, (5, 5), 0)
    _, thresh = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY_INV)

    kernel = np.ones((2, 2), np.uint8)
    thresh = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel)

    pixels = cv2.countNonZero(thresh)
    return pixels >= min_pixels, int(pixels)


def analyze_frequency_pdf(pdf_bytes: bytes, reference_month_name: str):
    """
    Analisa apenas a FREQUÊNCIA:
    - tenta localizar o mês esperado no OCR
    - verifica se há marca gráfica na área do residente
    - verifica se há marca gráfica na área do preceptor
    """

    result = {
        "status": "ok",
        "month_ok": True,
        "resident_signature_ok": True,
        "preceptor_signature_ok": True,
        "issues": [],
        "ocr_excerpt": "",
        "resident_pixels": 0,
        "preceptor_pixels": 0,
    }

    try:
        pages = convert_from_bytes(pdf_bytes, dpi=200, first_page=1, last_page=1)
    except Exception as e:
        result["status"] = "pendente"
        result["issues"].append(f"Falha ao converter PDF: {str(e)}")
        return result

    if not pages:
        result["status"] = "pendente"
        result["issues"].append("PDF sem páginas legíveis")
        return result

    first_page = pages[0]
    text = pytesseract.image_to_string(first_page, lang="por")
    result["ocr_excerpt"] = (text or "")[:500]

    # mês esperado vem de "Março 2026", por exemplo
    expected_month_name = normalize_ocr_text(reference_month_name.split()[0])
    if expected_month_name not in normalize_ocr_text(text):
        result["month_ok"] = False
        result["issues"].append("Mês do documento não confere com a competência esperada")

    img_cv = pil_to_cv2(first_page)

    # áreas iniciais de teste - provavelmente vamos ajustar depois
    resident_box = (0.05, 0.72, 0.48, 0.97)
    preceptor_box = (0.52, 0.72, 0.97, 0.97)

    resident_region = crop_relative_region(img_cv, resident_box)
    preceptor_region = crop_relative_region(img_cv, preceptor_box)

    resident_ok, resident_pixels = detect_handwritten_mark(resident_region)
    preceptor_ok, preceptor_pixels = detect_handwritten_mark(preceptor_region)

    result["resident_signature_ok"] = resident_ok
    result["preceptor_signature_ok"] = preceptor_ok
    result["resident_pixels"] = resident_pixels
    result["preceptor_pixels"] = preceptor_pixels

    if not resident_ok:
        result["issues"].append("Campo de assinatura/rubrica do residente aparenta estar vazio")

    if not preceptor_ok:
        result["issues"].append("Campo de assinatura/rubrica do preceptor aparenta estar vazio")

    if result["issues"]:
        result["status"] = "pendente"

    return result

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user_data: UserCreate):
    existing = await db.users.find_one({"email": user_data.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email já cadastrado")

    if user_data.role == "resident":
        if not user_data.year:
            raise HTTPException(status_code=400, detail="Year obrigatório para residente")

    elif user_data.role == "preceptor":
        # 👇 aqui decide se quer obrigar ou não
        if not user_data.cpf or not user_data.scenario:
            raise HTTPException(status_code=400, detail="CPF e cenário obrigatórios para preceptor")

    elif user_data.role == "admin":
        pass

    else:
        raise HTTPException(status_code=400, detail="Role inválido")

    user_dict = {
        "id": str(uuid.uuid4()),
        "email": user_data.email.lower(),
        "full_name": user_data.full_name,
        "program": user_data.program,
        "year": user_data.year,
        "password": hash_password(user_data.password),
        "role": user_data.role,
        "cpf": user_data.cpf,
        "scenario": user_data.scenario,
        "created_at": datetime.utcnow(),
        "is_active": True
    }

    await db.users.insert_one(user_dict)

    token = create_token(user_dict["id"], user_dict["role"])

    return TokenResponse(
        access_token=token,
        user=UserResponse(
            id=user_dict["id"],
            email=user_dict["email"],
            full_name=user_dict["full_name"],
            program=user_dict["program"],
            year=user_dict.get("year"),
            role=user_dict["role"],
            cpf=user_dict.get("cpf"),
            scenario=user_dict.get("scenario"),
            created_at=user_dict["created_at"],
            is_active=user_dict["is_active"]
        )
    )


@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email.lower()})
    if not user or not verify_password(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="Email ou senha inválidos")

    if not user.get("is_active", True):
        raise HTTPException(status_code=403, detail="Conta desativada")

    token = create_token(user["id"], user["role"])

    return TokenResponse(
    access_token=token,
    user=UserResponse(
        id=user["id"],
        email=user["email"],
        full_name=user["full_name"],
        program=user["program"],
        year=user.get("year"),
        role=user["role"],
        cpf=user.get("cpf"),
        scenario=user.get("scenario"),
        created_at=user["created_at"],
        is_active=user["is_active"]
    )
)


@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    return UserResponse(
    id=current_user["id"],
    email=current_user["email"],
    full_name=current_user["full_name"],
    program=current_user.get("program", ""),
    year=current_user.get("year"),
    role=current_user.get("role", "resident"),
    cpf=current_user.get("cpf"),
    scenario=current_user.get("scenario"),
    created_at=current_user["created_at"],
    is_active=current_user.get("is_active", True)
)


@api_router.get("/submissions/deadline-info")
async def get_deadline_info():
    return get_reference_month_info()

@api_router.post("/submissions/upload")
async def upload_pdf(
    file: Optional[UploadFile] = File(None),
    aulas_ministradas_file: Optional[UploadFile] = File(None),
    orientacao_trabalho_file: Optional[UploadFile] = File(None),
    current_user: dict = Depends(get_current_user)
):
    try:
        print("=== INICIO UPLOAD ===")
        print("file:", file.filename if file else None)
        print("aulas:", aulas_ministradas_file.filename if aulas_ministradas_file else None)
        print("orientacao:", orientacao_trabalho_file.filename if orientacao_trabalho_file else None)
        print("usuario:", current_user.get("email"))

        if not file or not aulas_ministradas_file or not orientacao_trabalho_file:
            raise HTTPException(
                status_code=400,
                detail="Todos os arquivos são obrigatórios"
            )

        deadline_info = get_reference_month_info()
        print("deadline_info:", deadline_info)

        files_to_validate = [
            ("Frequência", file),
            ("Aulas Ministradas", aulas_ministradas_file),
            ("Orientação de Trabalho", orientacao_trabalho_file),
        ]

        validated_files = {}

        for label, current_file in files_to_validate:
            print("Processando:", label, current_file.filename)

            if not current_file.filename.lower().endswith(".pdf"):
                raise HTTPException(
                    status_code=400,
                    detail=f"{label}: apenas arquivos PDF são aceitos"
                )

            content = await current_file.read()
            print(f"{label} tamanho:", len(content))

            if len(content) > 10 * 1024 * 1024:
                raise HTTPException(
                    status_code=400,
                    detail=f"{label}: arquivo muito grande. Máximo permitido: 10MB"
                )

            validated_files[label] = content

        print("Iniciando OCR da frequência")
        ocr_analysis = analyze_frequency_pdf(
            validated_files["Frequência"],
            deadline_info["reference_month_name"]
        )
        print("OCR resultado:", ocr_analysis)

        sanitized_name = sanitize_filename(current_user["full_name"])
        sanitized_program = sanitize_filename(current_user.get("program", "sem_programa"))
        month_name_sanitized = sanitize_filename(deadline_info["reference_month_name"])

        frequencia_folder = generate_frequencia_folder(
            current_user,
            deadline_info["reference_month"]
        )

        material_folder = generate_user_material_folder(
            current_user,
            deadline_info["reference_month"]
        )

        frequencia_filename = f"{sanitized_name}_{sanitized_program}_{month_name_sanitized}_frequencia.pdf"
        aulas_filename = f"{sanitized_name}_{month_name_sanitized}_aulas.pdf"
        orientacao_filename = f"{sanitized_name}_{month_name_sanitized}_orientacao.pdf"

        print("Iniciando upload Cloudinary da frequência")
        frequencia_cloud = upload_to_cloudinary(
            validated_files["Frequência"],
            frequencia_filename,
            frequencia_folder,
            f"{sanitized_name}_{sanitized_program}_frequencia"
        )
        print("Upload frequência OK")

        aulas_cloud = upload_to_cloudinary(
            validated_files["Aulas Ministradas"],
            aulas_filename,
            material_folder,
            "aulas"
        )
        print("Upload aulas OK")

        orientacao_cloud = upload_to_cloudinary(
            validated_files["Orientação de Trabalho"],
            orientacao_filename,
            material_folder,
            "orientacao"
        )
        print("Upload orientação OK")

        submission = {
            "id": str(uuid.uuid4()),
            "user_id": current_user["id"],
            "user_name": current_user["full_name"],
            "user_email": current_user["email"],
            "program": current_user.get("program", ""),
            "year": current_user.get("year") or "",
            "reference_month": deadline_info["reference_month"],
            "reference_month_name": deadline_info["reference_month_name"],
            "submitted_at": datetime.utcnow(),

            "file_path": frequencia_cloud["secure_url"],
            "file_name": frequencia_filename,

            "aulas_file_path": aulas_cloud["secure_url"],
            "aulas_file_name": aulas_filename,

            "orientacao_file_path": orientacao_cloud["secure_url"],
            "orientacao_file_name": orientacao_filename,

            "ocr_status": ocr_analysis["status"],
            "ocr_issues": ocr_analysis["issues"],
            "ocr_analysis": ocr_analysis,
        }

        print("Salvando submission no Mongo")
        await db.submissions.insert_one(submission)
        print("Submission salva com sucesso")
        print("=== FIM UPLOAD ===")

        return {
            "message": "Upload realizado com sucesso",
            "ocr_status": ocr_analysis["status"],
            "ocr_issues": ocr_analysis["issues"],
            "ocr_analysis": ocr_analysis
        }

    except HTTPException:
        raise

    except Exception as e:
        print("ERRO GERAL NO UPLOAD:", str(e))
        logger.exception("ERRO GERAL NO UPLOAD")
        raise HTTPException(
            status_code=500,
            detail=f"Erro interno no upload: {str(e)}"
        )

@api_router.get("/submissions/my-history", response_model=List[SubmissionResponse])
async def get_my_submissions(current_user: dict = Depends(get_current_user)):
    submissions = await db.submissions.find(
        {"user_id": current_user["id"]}
    ).sort("reference_month", -1).to_list(1000)

    return [
        SubmissionResponse(
            id=s["id"],
            user_id=s["user_id"],
            user_name=s["user_name"],
            user_email=s["user_email"],
            program=s.get("program", ""),
            year=s.get("year") or "",
            reference_month=s["reference_month"],
            reference_month_name=s["reference_month_name"],
            submitted_at=s["submitted_at"]
        )
        for s in submissions
    ]


@api_router.get("/submissions/check/{reference_month}")
async def check_submission(reference_month: str, current_user: dict = Depends(get_current_user)):
    existing = await db.submissions.find_one({
        "user_id": current_user["id"],
        "reference_month": reference_month
    })
    return {"submitted": existing is not None}


@api_router.get("/admin/submissions", response_model=List[SubmissionResponse])
async def get_all_submissions(
    reference_month: Optional[str] = None,
    program: Optional[str] = None,
    year: Optional[str] = None,
    admin_user: dict = Depends(get_admin_user)
):
    query = {}

    if reference_month:
        query["reference_month"] = reference_month
    if program:
        query["program"] = program
    if year:
        query["year"] = year

    submissions = await db.submissions.find(query).sort("user_name", 1).to_list(10000)

    return [
        SubmissionResponse(
            id=s["id"],
            user_id=s["user_id"],
            user_name=s["user_name"],
            user_email=s["user_email"],
            program=s.get("program", ""),
            year=s.get("year", ""),
            reference_month=s["reference_month"],
            reference_month_name=s["reference_month_name"],
            submitted_at=s["submitted_at"]
        )
        for s in submissions
    ]


@api_router.get("/admin/users", response_model=List[UserResponse])
async def get_all_users(admin_user: dict = Depends(get_admin_user)):
    users = await db.users.find({"role": "resident"}).sort("full_name", 1).to_list(10000)

    return [
        UserResponse(
            id=u["id"],
            email=u["email"],
            full_name=u["full_name"],
            program=u.get("program", ""),
            year=u.get("year", ""),
            role=u["role"],
            created_at=u["created_at"],
            is_active=u.get("is_active", True)
        )
        for u in users
    ]


@api_router.get("/admin/programs")
async def get_programs(admin_user: dict = Depends(get_admin_user)):
    programs = await db.users.distinct("program")
    return {"programs": [p for p in programs if p]}


@api_router.get("/admin/reference-months")
async def get_reference_months(admin_user: dict = Depends(get_admin_user)):
    months = await db.submissions.distinct("reference_month")
    months_sorted = sorted(months, reverse=True)

    result = []
    for m in months_sorted:
        year, month_num = m.split("-")
        month_name = f"{MONTH_NAMES_PT[int(month_num)]} {year}"
        result.append({"value": m, "label": month_name})

    return {"months": result}


@api_router.get("/admin/download/{submission_id}")
async def download_submission(submission_id: str, admin_user: dict = Depends(get_admin_user)):
    submission = await db.submissions.find_one({"id": submission_id})
    if not submission:
        raise HTTPException(status_code=404, detail="Documento não encontrado")

    response = requests.get(submission["file_path"])
    buffer = io.BytesIO(response.content)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={submission['file_name']}"}
    )


@api_router.get("/admin/export-excel")
async def export_excel(
    reference_month: Optional[str] = None,
    admin_user: dict = Depends(get_admin_user)
):
    query = {}
    if reference_month:
        query["reference_month"] = reference_month

    submissions = await db.submissions.find(query).sort("user_name", 1).to_list(10000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Envios"

    headers = ["Nome", "Email", "Programa", "Ano", "Mês de Referência", "Status", "Data e Hora do Envio"]
    ws.append(headers)

    red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")

    for s in submissions:
        status_excel = "Pendente OCR" if s.get("ocr_status") == "pendente" else "Enviado"

        ws.append([
            s["user_name"],
            s["user_email"],
            s.get("program", ""),
            s.get("year", ""),
            s["reference_month_name"],
            status_excel,
            s["submitted_at"].strftime("%d/%m/%Y %H:%M:%S")
        ])

        current_row = ws.max_row

        if s.get("ocr_status") == "pendente":
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).fill = red_fill

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"relatorio_envios_{reference_month or 'geral'}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/admin/export-zip")
async def export_zip(
    reference_month: str = Query(..., description="Reference month in format YYYY-MM"),
    admin_user: dict = Depends(get_admin_user)
):
    submissions = await db.submissions.find(
        {"reference_month": reference_month}
    ).sort("user_name", 1).to_list(10000)

    if not submissions:
        raise HTTPException(status_code=404, detail="Nenhum envio encontrado para este mês")

    buffer = io.BytesIO()

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for s in submissions:
            sanitized_name = sanitize_filename(s["user_name"])
            month_name_sanitized = sanitize_filename(s["reference_month_name"])
            new_filename = f"{sanitized_name}_{month_name_sanitized}.pdf"

            response = requests.get(s["file_path"])
            zf.writestr(new_filename, response.content)

    buffer.seek(0)

    year, month_num = reference_month.split("-")
    month_name = f"{MONTH_NAMES_PT[int(month_num)]}_{year}"

    return StreamingResponse(
        buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=PDFs_{month_name}.zip"}
    )
@api_router.get("/admin/export-zip-aulas")
async def export_zip_aulas(
    reference_month: str = Query(..., description="Reference month in format YYYY-MM"),
    admin_user: dict = Depends(get_admin_user)
):
    submissions = await db.submissions.find(
        {"reference_month": reference_month}
    ).sort("user_name", 1).to_list(10000)

    if not submissions:
        raise HTTPException(status_code=404, detail="Nenhum envio encontrado para este mês")

    buffer = io.BytesIO()

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for s in submissions:
            if not s.get("aulas_file_path"):
                continue

            response = requests.get(s["aulas_file_path"])
            sanitized_name = sanitize_filename(s["user_name"])
            month_name_sanitized = sanitize_filename(s["reference_month_name"])
            new_filename = f"{sanitized_name}_{month_name_sanitized}_aulas.pdf"
            zf.writestr(new_filename, response.content)

    buffer.seek(0)

    year, month_num = reference_month.split("-")
    month_name = f"{MONTH_NAMES_PT[int(month_num)]}_{year}"

    return StreamingResponse(
        buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=Aulas_{month_name}.zip"}
    )

@api_router.get("/admin/export-zip-orientacao")
async def export_zip_orientacao(
    reference_month: str = Query(..., description="Reference month in format YYYY-MM"),
    admin_user: dict = Depends(get_admin_user)
):
    submissions = await db.submissions.find(
        {"reference_month": reference_month}
    ).sort("user_name", 1).to_list(10000)

    if not submissions:
        raise HTTPException(status_code=404, detail="Nenhum envio encontrado para este mês")

    buffer = io.BytesIO()

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for s in submissions:
            if not s.get("orientacao_file_path"):
                continue

            sanitized_name = sanitize_filename(s["user_name"])
            month_name_sanitized = sanitize_filename(s["reference_month_name"])
            new_filename = f"{sanitized_name}_{month_name_sanitized}_orientacao.pdf"

            response = requests.get(s["orientacao_file_path"])
            zf.writestr(new_filename, response.content)

    buffer.seek(0)

    year, month_num = reference_month.split("-")
    month_name = f"{MONTH_NAMES_PT[int(month_num)]}_{year}"

    return StreamingResponse(
        buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=Orientacao_{month_name}.zip"}
    )

@api_router.get("/config")
async def get_system_config():
    config = await db.system_config.find_one({"id": "system_config"})
    if not config:
        return SystemConfig().dict()
    return config


@api_router.put("/admin/config")
async def update_system_config(
    config_update: SystemConfigUpdate,
    admin_user: dict = Depends(get_admin_user)
):
    update_data = {k: v for k, v in config_update.dict().items() if v is not None}
    update_data["updated_at"] = datetime.utcnow()

    await db.system_config.update_one(
        {"id": "system_config"},
        {"$set": update_data},
        upsert=True
    )

    config = await db.system_config.find_one({"id": "system_config"})
    return config


@api_router.get("/admin/stats")
async def get_stats(admin_user: dict = Depends(get_admin_user)):
    total_users = await db.users.count_documents({"role": "resident"})
    total_preceptors = await db.users.count_documents({"role": "preceptor"})
    total_submissions = await db.submissions.count_documents({})

    preceptors = await db.users.find({"role": "preceptor"}).to_list(None)
    preceptor_ids = [p["id"] for p in preceptors]

    total_preceptor_submissions = await db.submissions.count_documents({
        "user_id": {"$in": preceptor_ids}
    })

    deadline_info = get_reference_month_info()
    current_month_submissions = await db.submissions.count_documents({
        "reference_month": deadline_info["reference_month"]
    })

    pipeline = [
        {"$group": {"_id": "$program", "count": {"$sum": 1}}}
    ]
    by_program = await db.submissions.aggregate(pipeline).to_list(100)

    return {
        "total_users": total_users,
        "total_preceptors": total_preceptors,
        "total_submissions": total_submissions,
        "total_preceptor_submissions": total_preceptor_submissions,
        "current_month": deadline_info["reference_month_name"],
        "current_month_submissions": current_month_submissions,
        "is_within_deadline": deadline_info["is_within_deadline"],
        "by_program": by_program
    }
@api_router.get("/")
async def root():
    return {"message": "COREME UEPA API", "status": "online"}


@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.utcnow().isoformat()}


app.include_router(api_router)

@app.on_event("startup")
async def startup_db_client():
    try:
        admin = await db.users.find_one({"email": "admin@uepa.br"})

        if not admin:
            admin_user = {
                "id": str(uuid.uuid4()),
                "email": "admin@uepa.br",
                "full_name": "Administrador COREME",
                "program": "Administração",
                "year": "N/A",
                "password": hash_password("admin123"),
                "role": "admin",
                "created_at": datetime.utcnow(),
                "is_active": True
            }
            await db.users.insert_one(admin_user)
            logger.info("Default admin user created: admin@uepa.br / admin123")

        await db.users.create_index("email", unique=True)
        await db.users.create_index("id")
        await db.submissions.create_index([("user_id", 1), ("reference_month", 1)])
        await db.submissions.create_index("reference_month")

        print("Banco conectado com sucesso")

    except Exception as e:
        print("Erro ao conectar no banco:", e)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
